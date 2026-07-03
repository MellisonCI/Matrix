"""
One-time ETL: parse the Bank Capabilities Matrix and Bank Product Matrix
Excel files into the normalized schema in supabase-schema.sql.

This is a throwaway migration script, not a maintained part of the app.
Run modes:
  python import_matrices.py                 -> dry run: parse + print validation report only
  python import_matrices.py --load           -> parse + load into Supabase (requires DATABASE_URL env var)

Design notes (see plan for full rationale):
- ws.dimensions / max_row / max_column are unreliable in these files (report
  bogus XFB-range extents), so real extents are found by scanning column A /
  row 1 for the last non-null cell.
- Row hierarchy is encoded via cell.alignment.indent (0/1/2). Classification
  is driven by content, not cell fill -- fill='solid' is used both for
  genuine section headers AND, inconsistently, for real leaf features the
  sheet author wanted visually emphasized (confirmed exceptions: "Alerts" >
  "Statement Available/Current Statement Closes" is indent 0 + solid fill
  with real data; "Help and Services" > "Embedded Functionality/Information"
  is indent 1 + solid fill with real data). The actual rule:
    * indent == 0, row is empty or matches the junk-progression pattern below
      -> subcategory header
    * indent == 0, row has real data -> root feature, filed under whatever
      subcategory was already current (parent_feature_id NULL)
    * indent > 0 -> feature nested under the last-seen feature one level up
- Some header rows contain a bogus arithmetic-progression of tiny floats
  (step ~= 1/(n-1)) instead of blanks -- an Excel artifact, detected and
  discarded rather than treated as data.
- Only non-blank cells produce a value row; a blank cell means "no data" and
  is simply not inserted (no is_present=false row), keeping the values
  tables lean.
"""

import argparse
import os
import sys
from dataclasses import dataclass, field

import openpyxl

CAPABILITIES_FILE = "Bank Capabilities Matrix Q1 2026.xlsx"
PRODUCT_FILE = "Bank Product Matrix Q1 2026.xlsx"

QUARTER_LABEL = "Q1 2026"
QUARTER_YEAR = 2026
QUARTER_NUMBER = 1

FIRM_NAMES = [
    "Ally", "American Express", "Bank of America", "Capital One",
    "Charles Schwab", "Chase", "Chime", "Citi", "Citizens Bank", "Discover",
    "Fifth Third", "PNC", "TD Bank", "Truist", "U.S. Bank", "USAA", "Wells Fargo",
]

CAPABILITY_CATEGORY_SHEETS = [
    "Public Site",
    "Authentication and Security",
    "Account Information",
    "Transaction History",
    "Alerts",
    "Statements",
    "Transfers and Bill Pay",
    "Account Maintenance",
    "Help and Services",
]

PRODUCT_CATEGORY_SHEETS = [
    "Checking Accounts",
    "Savings & Money Market Accounts",
    "Certificates of Deposit",
]

BULLET = "•"

# Hand-built mapping of every raw product-sheet column header to (firm, product name).
# Built by manually reviewing all 95 headers across the three product sheets --
# roughly half have no newline separating firm from product name (e.g. "Citigold"),
# so this can't be parsed generically. The import script asserts every header in
# the sheet is present here and fails loudly if not.
PRODUCT_HEADER_MAP = {
    # Checking Accounts
    "Ally\nSpending Account": ("Ally", "Spending Account"),
    "American Express Rewards Checking": ("American Express", "Rewards Checking"),
    "Bank of America\nAdvantage SafeBalance": ("Bank of America", "Advantage SafeBalance"),
    "Bank of America\nAdvantage Plus": ("Bank of America", "Advantage Plus"),
    "Bank of America\nAdvantage Relationship": ("Bank of America", "Advantage Relationship"),
    "Capital One\n360 Checking": ("Capital One", "360 Checking"),
    "Charles Schwab\nBank Investor Checking": ("Charles Schwab", "Bank Investor Checking"),
    "Chase\nSecure Banking": ("Chase", "Secure Banking"),
    "Chase\nTotal Checking": ("Chase", "Total Checking"),
    "Chase\nPremier Plus Checking": ("Chase", "Premier Plus Checking"),
    "Chime Online Checking": ("Chime", "Online Checking"),
    "Citi\nAccess Checking": ("Citi", "Access Checking"),
    "Citi\nRegular Checking": ("Citi", "Regular Checking"),
    "Citi Priority": ("Citi", "Priority"),
    "Citigold": ("Citi", "Citigold"),
    "Citizens Bank\nOne Deposit Checking": ("Citizens Bank", "One Deposit Checking"),
    "Citizens EverValue Checking": ("Citizens Bank", "EverValue Checking"),
    "Citizens Bank Quest Checking": ("Citizens Bank", "Quest Checking"),
    "Fifth Third\nExpress Checking": ("Fifth Third", "Express Checking"),
    "Fifth Third Momentum Checking": ("Fifth Third", "Momentum Checking"),
    "Fifth Third\nPreferred Checking": ("Fifth Third", "Preferred Checking"),
    "PNC Simple Checking": ("PNC", "Simple Checking"),
    "PNC\nVirtual Wallet": ("PNC", "Virtual Wallet"),
    "PNC\nVirtual Wallet \nPerformance Select": ("PNC", "Virtual Wallet Performance Select"),
    "TD Bank Essential Checking": ("TD Bank", "Essential Checking"),
    "TD Bank Complete Checking": ("TD Bank", "Complete Checking"),
    "TD Bank\nBeyond Checking": ("TD Bank", "Beyond Checking"),
    "Truist\nOne Checking": ("Truist", "One Checking"),
    "Truist Confidence": ("Truist", "Confidence"),
    "Truist Marquee Checking": ("Truist", "Marquee Checking"),
    "U.S. Bank\nSafe Debit": ("U.S. Bank", "Safe Debit"),
    "U.S. Bank Smartly Checking": ("U.S. Bank", "Smartly Checking"),
    "USAA\nClassic Checking": ("USAA", "Classic Checking"),
    "Wells Fargo Clear Access Banking": ("Wells Fargo", "Clear Access Banking"),
    "Wells Fargo\nEveryday Checking": ("Wells Fargo", "Everyday Checking"),
    "Wells Fargo Prime Checking": ("Wells Fargo", "Prime Checking"),
    "Wells Fargo Premier Checking": ("Wells Fargo", "Premier Checking"),
    # Savings & Money Market Accounts
    "Ally \nSavings": ("Ally", "Savings"),
    "Ally \nMoney Market": ("Ally", "Money Market"),
    "American Express High Yield Savings": ("American Express", "High Yield Savings"),
    "Bank of America\nAdvantage Savings": ("Bank of America", "Advantage Savings"),
    "Capital One\n360 Performance Savings": ("Capital One", "360 Performance Savings"),
    "Charles Schwab\nBank Investor Savings": ("Charles Schwab", "Bank Investor Savings"),
    "Chase\nSavings": ("Chase", "Savings"),
    "Chase\nPremier Savings": ("Chase", "Premier Savings"),
    "Chime High-Yield Savings": ("Chime", "High-Yield Savings"),
    "Citi\nSavings": ("Citi", "Savings"),
    "Citizens Bank\nOne Deposit Savings": ("Citizens Bank", "One Deposit Savings"),
    "Citizens Bank Quest Savings": ("Citizens Bank", "Quest Savings"),
    "Citizens Bank\nPersonal Money Market": ("Citizens Bank", "Personal Money Market"),
    "Citizens Bank Quest Money Market": ("Citizens Bank", "Quest Money Market"),
    "Fifth Third\nMomentum Savings": ("Fifth Third", "Momentum Savings"),
    "Fifth Third\nRelationship Money Market": ("Fifth Third", "Relationship Money Market"),
    "PNC \nStandard Savings": ("PNC", "Standard Savings"),
    "PNC \nPremiere Money Market": ("PNC", "Premiere Money Market"),
    "TD Signature Savings": ("TD Bank", "Signature Savings"),
    "TD Bank\nSimple Savings": ("TD Bank", "Simple Savings"),
    "Truist Confidence Savings": ("Truist", "Confidence Savings"),
    "Truist One Savings": ("Truist", "One Savings"),
    "Truist One Money Market": ("Truist", "One Money Market"),
    "U.S. Bank Smartly Savings": ("U.S. Bank", "Smartly Savings"),
    "U.S. Bank\nElite Money Market": ("U.S. Bank", "Elite Money Market"),
    "USAA\nSavings": ("USAA", "Savings"),
    "USAA\nPerformance First Savings": ("USAA", "Performance First Savings"),
    "Wells Fargo\nWay2Save": ("Wells Fargo", "Way2Save"),
    "Wells Fargo\nPlatinum Savings": ("Wells Fargo", "Platinum Savings"),
    # Certificates of Deposit
    "Ally \nHigh Yield CD": ("Ally", "High Yield CD"),
    "Ally \nRaise Your Rate CD": ("Ally", "Raise Your Rate CD"),
    "Ally \nNo Penalty CD": ("Ally", "No Penalty CD"),
    "American Express CD": ("American Express", "CD"),
    "Bank of America\nFeatured CD": ("Bank of America", "Featured CD"),
    "Bank of America\nFixed Term CD": ("Bank of America", "Fixed Term CD"),
    "Bank of America\nFlexible CD": ("Bank of America", "Flexible CD"),
    "Capital One\n360 CD": ("Capital One", "360 CD"),
    "Charles Schwab Certificates of Deposit": ("Charles Schwab", "Certificates of Deposit"),
    "Chase\nCertificates of Deposit": ("Chase", "Certificates of Deposit"),
    "Citi\nFixed Rate CD": ("Citi", "Fixed Rate CD"),
    "Citi\nStep Up CD": ("Citi", "Step Up CD"),
    "Citi\nNo Penalty CD": ("Citi", "No Penalty CD"),
    "Citizens Bank\nCertificates of Deposit": ("Citizens Bank", "Certificates of Deposit"),
    "Fifth Third \nStandard CD": ("Fifth Third", "Standard CD"),
    "Fifth Third \nFeatured CD": ("Fifth Third", "Featured CD"),
    "PNC \nFixed Rate CD": ("PNC", "Fixed Rate CD"),
    "TD Bank\nChoice Promotional CD": ("TD Bank", "Choice Promotional CD"),
    "TD Bank\nNo-Catch CD": ("TD Bank", "No-Catch CD"),
    "Truist Certificates of Deposit": ("Truist", "Certificates of Deposit"),
    "U.S. Bank\nStandard CD": ("U.S. Bank", "Standard CD"),
    "U.S. Bank\nCD Special": ("U.S. Bank", "CD Special"),
    "U.S. Bank\nStep Up CD": ("U.S. Bank", "Step Up CD"),
    "U.S. Bank\nTrade Up CD": ("U.S. Bank", "Trade Up CD"),
    "USAA\nStandard CD": ("USAA", "Standard CD"),
    "Wells Fargo\nFixed Rate CD": ("Wells Fargo", "Fixed Rate CD"),
}


# ------------------------------------------------------------------
# Data model (in-memory, mirrors the DB schema before it's loaded)
# ------------------------------------------------------------------

@dataclass
class Feature:
    name: str
    depth: int
    parent: "Feature | None"
    subcategory_name: str
    value_type: str = "boolean"
    values: dict = field(default_factory=dict)  # column_key -> parsed value dict
    children: list = field(default_factory=list)
    db_id: str | None = None  # filled in by load_into_supabase once inserted


@dataclass
class ParsedSheet:
    category_name: str
    column_keys: list  # firm names, or product header strings
    subcategories: list  # list of subcategory names, in order
    features: list  # flat list of Feature, in row order
    warnings: list = field(default_factory=list)
    adoption_mismatches: list = field(default_factory=list)
    discarded_artifact_rows: list = field(default_factory=list)


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def real_last_row(ws, col=1, scan_limit=2000):
    last = 0
    for r in range(1, scan_limit + 1):
        if ws.cell(row=r, column=col).value not in (None, ""):
            last = r
    return last


def real_last_col(ws, row=1, scan_limit=200):
    last = 0
    for c in range(1, scan_limit + 1):
        if ws.cell(row=row, column=c).value not in (None, ""):
            last = c
    return last


def is_junk_progression(values, unit=1.0 / 30, tolerance=0.001):
    """
    Detects the leftover-Excel-artifact rows: tiny floats that sit on a 1/30 grid
    and strictly increase left-to-right (occasionally skipping a step, apparently
    from columns deleted after the formula was hardcoded to values). Real numeric
    data (dollar limits, month counts, APYs) never satisfies "strictly increasing
    across every populated column in sheet order", so requiring strict monotonicity
    -- not just "on the 1/30 grid" -- is what actually distinguishes junk from data
    (small APY-type decimals also sit near k=0 on this grid but repeat/aren't increasing).
    """
    steps = []
    for v in values:
        if v is None:
            continue
        if not isinstance(v, (int, float)):
            return False
        k = v / unit
        nearest = round(k)
        if abs(v - nearest * unit) > tolerance:
            return False
        steps.append(nearest)
    if len(steps) < 3:
        return False
    return all(steps[i] < steps[i + 1] for i in range(len(steps) - 1))


def parse_cell(raw):
    """Returns dict(raw_text, is_present, is_not_applicable, numeric_value, detail)."""
    empty = dict(raw_text=None, is_present=None, is_not_applicable=False,
                 numeric_value=None, detail=None)
    if raw is None:
        return empty
    if isinstance(raw, (int, float)):
        return dict(raw_text=str(raw), is_present=None, is_not_applicable=False,
                    numeric_value=float(raw), detail=None)
    text = str(raw).strip()
    if text == "":
        return empty
    if text.upper() in ("N/A", "NA"):
        return dict(raw_text=text, is_present=None, is_not_applicable=True,
                    numeric_value=None, detail=None)
    if BULLET in text:
        detail = text.replace(BULLET, "").strip()
        if detail.startswith("(") and detail.endswith(")"):
            detail = detail[1:-1].strip()
        return dict(raw_text=text, is_present=True, is_not_applicable=False,
                    numeric_value=None, detail=detail or None)
    cleaned = text.replace(",", "").replace("$", "")
    try:
        num = float(cleaned)
        return dict(raw_text=text, is_present=None, is_not_applicable=False,
                    numeric_value=num, detail=None)
    except ValueError:
        pass
    return dict(raw_text=text, is_present=None, is_not_applicable=False,
                numeric_value=None, detail=None)


def infer_value_type(values):
    """Majority vote across non-blank, non-N/A cell parses for one feature row."""
    bullet_count = sum(1 for v in values if v["is_present"] is True)
    numeric_count = sum(1 for v in values if v["numeric_value"] is not None)
    text_count = sum(
        1 for v in values
        if v["raw_text"] is not None and v["is_present"] is None
        and v["numeric_value"] is None and not v["is_not_applicable"]
    )
    counts = {"boolean": bullet_count, "numeric": numeric_count, "text": text_count}
    best = max(counts, key=counts.get)
    if counts[best] == 0:
        return "boolean"
    return best


# ------------------------------------------------------------------
# Sheet parsing
# ------------------------------------------------------------------

def parse_capability_sheet(ws, sheet_name):
    last_col = real_last_col(ws)
    last_row = real_last_row(ws)
    firm_names = []
    for c in range(3, last_col + 1):  # column 2 is "Industry Adoption"
        v = ws.cell(row=1, column=c).value
        firm_names.append(str(v).strip() if v is not None else None)

    return _parse_hierarchical_sheet(
        ws, sheet_name, first_data_col=3, last_col=last_col, last_row=last_row,
        column_keys=firm_names, industry_adoption_col=2,
    )


def parse_product_sheet(ws, sheet_name):
    last_col = real_last_col(ws)
    last_row = real_last_row(ws)
    raw_headers = []
    for c in range(2, last_col + 1):
        v = ws.cell(row=1, column=c).value
        raw_headers.append(v)

    unmapped = [h for h in raw_headers if h not in PRODUCT_HEADER_MAP]
    if unmapped:
        raise SystemExit(
            f"Unmapped product headers in sheet '{sheet_name}': {unmapped!r}\n"
            "Add these to PRODUCT_HEADER_MAP before importing."
        )
    column_keys = [PRODUCT_HEADER_MAP[h] for h in raw_headers]  # (firm, product) tuples

    return _parse_hierarchical_sheet(
        ws, sheet_name, first_data_col=2, last_col=last_col, last_row=last_row,
        column_keys=column_keys, industry_adoption_col=None,
    )


def _parse_hierarchical_sheet(ws, sheet_name, first_data_col, last_col, last_row,
                               column_keys, industry_adoption_col):
    parsed = ParsedSheet(category_name=sheet_name, column_keys=column_keys,
                          subcategories=[], features=[])
    stack = {}  # depth -> Feature (last-seen feature at that depth)
    current_subcategory = None

    for r in range(2, last_row + 1):
        label_cell = ws.cell(row=r, column=1)
        label = label_cell.value
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()
        indent = int(label_cell.alignment.indent or 0) if label_cell.alignment else 0

        row_values_raw = [ws.cell(row=r, column=c).value for c in range(first_data_col, last_col + 1)]
        is_junk = is_junk_progression(row_values_raw)
        is_empty = all(v is None for v in row_values_raw)

        # Classification is driven by content, not cell fill: a fill='solid' style is
        # used both for genuine section headers AND, inconsistently, for real leaf
        # features the sheet author wanted visually emphasized (confirmed exception:
        # "Alerts" > "Statement Available/Current Statement Closes" is indent 0, solid
        # fill, and has real per-firm data). So indent 0 is a header only when its row
        # is actually empty/junk; otherwise it's a real root-level feature that stays
        # under whatever subcategory was already current.
        if indent == 0 and (is_empty or is_junk):
            current_subcategory = label
            parsed.subcategories.append(label)
            stack = {}
            if is_junk:
                parsed.discarded_artifact_rows.append((r, label))
            continue

        if current_subcategory is None:
            current_subcategory = "(ungrouped)"
            parsed.subcategories.append(current_subcategory)

        parent = stack.get(indent - 1) if indent > 0 else None
        feature = Feature(name=label, depth=indent, parent=parent,
                           subcategory_name=current_subcategory)
        if parent is not None:
            parent.children.append(feature)

        parsed_values = [parse_cell(v) for v in row_values_raw]
        if is_junk:
            parsed.discarded_artifact_rows.append((r, label))
            parsed_values = [parse_cell(None) for _ in row_values_raw]

        for key, pv in zip(column_keys, parsed_values):
            if pv["raw_text"] is not None:
                feature.values[key] = pv

        feature.value_type = infer_value_type(parsed_values)

        if industry_adoption_col is not None:
            stated = ws.cell(row=r, column=industry_adoption_col).value
            if isinstance(stated, (int, float)) and 0 <= stated <= 1 and feature.value_type == "boolean":
                present_count = sum(1 for v in parsed_values if v["is_present"])
                total = sum(1 for v in parsed_values if not v["is_not_applicable"])
                computed = present_count / total if total else 0
                if abs(computed - stated) > 0.03:
                    parsed.adoption_mismatches.append(
                        (r, label, round(stated, 3), round(computed, 3))
                    )

        stack[indent] = feature
        stack = {d: f for d, f in stack.items() if d <= indent}
        parsed.features.append(feature)

    return parsed


# ------------------------------------------------------------------
# Report
# ------------------------------------------------------------------

def print_report(parsed: ParsedSheet):
    n_features = len(parsed.features)
    n_nested = sum(1 for f in parsed.features if f.parent is not None)
    n_values = sum(len(f.values) for f in parsed.features)
    print(f"\n=== {parsed.category_name} ===")
    print(f"  subcategories: {len(parsed.subcategories)} -> {parsed.subcategories}")
    print(f"  features: {n_features} (nested under a parent: {n_nested})")
    print(f"  values inserted: {n_values}")
    type_counts = {}
    for f in parsed.features:
        type_counts[f.value_type] = type_counts.get(f.value_type, 0) + 1
    print(f"  value_type distribution: {type_counts}")
    if parsed.discarded_artifact_rows:
        print(f"  discarded arithmetic-progression artifact rows: {parsed.discarded_artifact_rows}")
    if parsed.warnings:
        for w in parsed.warnings:
            print(f"  WARNING: {w}")
    if parsed.adoption_mismatches:
        print(f"  Industry Adoption mismatches (row, label, stated, computed):")
        for m in parsed.adoption_mismatches:
            print(f"    {m}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--load", action="store_true", help="Load into Supabase (requires DATABASE_URL)")
    args = ap.parse_args()

    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cap_path = os.path.join(base, CAPABILITIES_FILE)
    prod_path = os.path.join(base, PRODUCT_FILE)

    print("Loading workbooks (this can take a little while for the 13MB capabilities file)...")
    cap_wb = openpyxl.load_workbook(cap_path, data_only=True)
    prod_wb = openpyxl.load_workbook(prod_path, data_only=True)

    cap_parsed = []
    for sheet_name in CAPABILITY_CATEGORY_SHEETS:
        parsed = parse_capability_sheet(cap_wb[sheet_name], sheet_name)
        cap_parsed.append(parsed)
        print_report(parsed)

    prod_parsed = []
    for sheet_name in PRODUCT_CATEGORY_SHEETS:
        parsed = parse_product_sheet(prod_wb[sheet_name], sheet_name)
        prod_parsed.append(parsed)
        print_report(parsed)

    print("\n=== Product header -> (firm, product) map used ===")
    for h, (firm, product) in PRODUCT_HEADER_MAP.items():
        print(f"  {h!r} -> firm={firm!r} product={product!r}")

    if args.load:
        from load_into_supabase import load_all
        load_all(cap_parsed, prod_parsed)
    else:
        print("\nDry run complete. Re-run with --load (and DATABASE_URL set) to write to Supabase.")


if __name__ == "__main__":
    main()
